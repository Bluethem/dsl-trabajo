from flask import Flask, render_template, request, jsonify, send_file, session
from jinja2 import Environment, BaseLoader, TemplateSyntaxError, UndefinedError
from markupsafe import Markup
import json
import os
import io
import sqlite3
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = "dsl-reportes-secret-2025"

DB_PATH = os.path.join(os.path.dirname(__file__), "reports", "templates.db")

# ── SQLite init ──────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = get_db()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS saved_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            template TEXT NOT NULL,
            data TEXT NOT NULL DEFAULT '{}',
            version INTEGER NOT NULL DEFAULT 1,
            visibility TEXT NOT NULL DEFAULT 'public',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS scheduled_reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            template_id INTEGER,
            frequency TEXT NOT NULL DEFAULT 'daily',
            active INTEGER NOT NULL DEFAULT 1,
            last_run TIMESTAMP,
            next_run TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS report_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            schedule_id INTEGER NOT NULL,
            schedule_name TEXT,
            status TEXT NOT NULL DEFAULT 'ok',
            ran_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # Migrate: add visibility column if it doesn't exist yet
    try:
        conn.execute("ALTER TABLE saved_templates ADD COLUMN visibility TEXT NOT NULL DEFAULT 'public'")
    except Exception:
        pass
    conn.commit()
    conn.close()

init_db()

# ── Jinja2 DSL environment ───────────────────────────────────────────────────

user_env = Environment(loader=BaseLoader())

# ── Custom DSL filters ───────────────────────────────────────────────────────

def currency_filter(value, symbol="S/"):
    try:
        return f"{symbol} {float(value):,.2f}"
    except (ValueError, TypeError):
        return value

def upper_first(value):
    return str(value).capitalize()

def badge(value, thresholds=None):
    if thresholds is None:
        thresholds = {"danger": 30, "warning": 70}
    try:
        v = float(value)
        if v < thresholds["danger"]:
            return "badge-danger"
        elif v < thresholds["warning"]:
            return "badge-warning"
        return "badge-success"
    except (ValueError, TypeError):
        return "badge-default"

def percent_filter(value, decimals=1):
    try:
        return f"{float(value):.{decimals}f}%"
    except (ValueError, TypeError):
        return value

def tojson_filter(value, indent=None):
    return Markup(json.dumps(value, ensure_ascii=False, indent=indent))

def chart_filter(config):
    """
    Convierte un dict de configuración de Chart.js en un <canvas> listo
    para ser inicializado por el frontend.
    Uso en plantilla: {{ chart_config | chart }}
    Donde chart_config es un dict con keys: type, labels, datasets, title
    """
    chart_json = json.dumps(config, ensure_ascii=False)
    return Markup(
        f'<canvas class="dsl-chart" '
        f'data-chart=\'{chart_json}\' '
        f'style="max-width:100%;max-height:320px;margin:12px 0"></canvas>'
    )

user_env.filters["currency"]   = currency_filter
user_env.filters["upper_first"] = upper_first
user_env.filters["badge"]      = badge
user_env.filters["percent"]    = percent_filter
user_env.filters["tojson"]     = tojson_filter
user_env.filters["chart"]      = chart_filter

# ── Example templates ────────────────────────────────────────────────────────

EXAMPLE_TEMPLATES = {
    "ventas": {
        "name": "Reporte de Ventas",
        "template": """\
<h1>{{ titulo }}</h1>
<p>Periodo: {{ periodo }}</p>
<table>
  <thead>
    <tr><th>Producto</th><th>Cantidad</th><th>Total</th></tr>
  </thead>
  <tbody>
    {% for item in ventas %}
    <tr>
      <td>{{ item.producto | upper_first }}</td>
      <td>{{ item.cantidad }}</td>
      <td>{{ item.total | currency }}</td>
    </tr>
    {% endfor %}
  </tbody>
</table>
<p><strong>Total general: {{ ventas | sum(attribute='total') | currency }}</strong></p>
""",
        "data": json.dumps({
            "titulo": "Reporte Mensual",
            "periodo": "Junio 2025",
            "ventas": [
                {"producto": "laptop",  "cantidad": 5,  "total": 15000},
                {"producto": "mouse",   "cantidad": 20, "total": 600},
                {"producto": "teclado", "cantidad": 12, "total": 1800}
            ]
        }, indent=2)
    },
    "empleados": {
        "name": "Reporte de Empleados",
        "template": """\
<h1>{{ empresa }}</h1>
<h2>Personal Activo</h2>
{% for emp in empleados %}
<div class="card">
  <strong>{{ emp.nombre }}</strong> — {{ emp.cargo }}
  <span class="{{ emp.rendimiento | badge }}">{{ emp.rendimiento | percent }}</span>
</div>
{% endfor %}
<p>Total empleados: {{ empleados | length }}</p>
""",
        "data": json.dumps({
            "empresa": "Soluciones Técnicas Andinas S.A.C.",
            "empleados": [
                {"nombre": "Ana Torres",   "cargo": "Desarrolladora", "rendimiento": 92},
                {"nombre": "Luis Quispe",  "cargo": "Diseñador",      "rendimiento": 65},
                {"nombre": "María Chávez", "cargo": "QA",             "rendimiento": 25}
            ]
        }, indent=2)
    },
    "ventas_chart": {
        "name": "Ventas con Gráfico",
        "template": """\
<h1>{{ titulo }}</h1>
<p>Periodo: {{ periodo }}</p>

{% set chart_config = {
  "type": "bar",
  "title": "Ventas por Producto",
  "labels": ventas | map(attribute="producto") | list,
  "datasets": [{
    "label": "Total (S/)",
    "data": ventas | map(attribute="total") | list,
    "backgroundColor": ["#7c6af7","#56cfb2","#f59e0b"]
  }]
} %}
{{ chart_config | chart }}

<table>
  <thead>
    <tr><th>Producto</th><th>Cantidad</th><th>Total</th></tr>
  </thead>
  <tbody>
    {% for item in ventas %}
    <tr>
      <td>{{ item.producto | upper_first }}</td>
      <td>{{ item.cantidad }}</td>
      <td>{{ item.total | currency }}</td>
    </tr>
    {% endfor %}
  </tbody>
</table>
<p><strong>Total general: {{ ventas | sum(attribute='total') | currency }}</strong></p>
""",
        "data": json.dumps({
            "titulo": "Reporte de Ventas — Junio 2025",
            "periodo": "Junio 2025",
            "ventas": [
                {"producto": "Laptop",  "cantidad": 5,  "total": 15000},
                {"producto": "Mouse",   "cantidad": 20, "total": 600},
                {"producto": "Teclado", "cantidad": 12, "total": 1800}
            ]
        }, indent=2)
    },
    "permisos": {
        "name": "Reporte con Permisos",
        "template": """\
<h1>Reporte Empresarial</h1>
<p>Usuario: <strong>{{ _user }}</strong> — Rol: <span class="{{ _role == 'admin' | string | lower == 'true' and 'badge-success' or 'badge-warning' }}">{{ _role }}</span></p>

<h2>Resumen General</h2>
<p>Visible para todos los roles.</p>
<table>
  <thead><tr><th>Área</th><th>Ventas</th></tr></thead>
  <tbody>
    {% for item in areas %}
    <tr><td>{{ item.area }}</td><td>{{ item.ventas | currency }}</td></tr>
    {% endfor %}
  </tbody>
</table>

{% if _role == 'admin' %}
<div class="admin-section">
  <h2>Datos Confidenciales (solo Admin)</h2>
  <p>Margen de utilidad: <strong>{{ margen | percent }}</strong></p>
  <p>Costo operativo: <strong>{{ costo_op | currency }}</strong></p>
  <div class="card">
    <strong>Acceso restringido:</strong> Esta sección no es visible para el rol <em>viewer</em>.
  </div>
</div>
{% else %}
<div class="card">
  <span class="badge-warning">Acceso limitado</span>
  Inicia sesión como <strong>admin</strong> para ver los datos confidenciales.
</div>
{% endif %}
""",
        "data": json.dumps({
            "areas": [
                {"area": "Norte",  "ventas": 45000},
                {"area": "Sur",    "ventas": 32000},
                {"area": "Centro", "ventas": 61000}
            ],
            "margen": 23.5,
            "costo_op": 87500
        }, indent=2)
    },
    "demo_extrema": {
        "name": "SVG Video Lab",
        "template": """\
<style>
  .bi-hero { background: linear-gradient(135deg, #111827, #243b55); color: white; padding: 18px; border-radius: 10px; margin-bottom: 14px; }
  .bi-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 10px; margin: 12px 0; }
  .bi-kpi { background: #f8fafc; border: 1px solid #e5e7eb; border-radius: 8px; padding: 12px; }
  .bi-kpi small { color: #64748b; display: block; margin-bottom: 4px; }
  .bi-kpi strong { font-size: 1.25rem; color: #111827; }
  .pivot td, .pivot th { text-align: right; }
  .pivot td:first-child, .pivot th:first-child { text-align: left; font-weight: 700; }
  .heat-low { background: #fee2e2 !important; }
  .heat-mid { background: #fef3c7 !important; }
  .heat-high { background: #dcfce7 !important; }
  .svg-panel { background: #020617; color: white; border-radius: 12px; padding: 14px; margin: 14px 0; box-shadow: inset 0 0 0 1px rgba(148,163,184,.2); }
  .svg-panel h2 { color: white; margin-top: 0; }
  .scene-caption { color: #94a3b8; font-size: .82rem; margin: 8px 0 0; }
  .video-stage { display: grid; grid-template-columns: 1fr; gap: 12px; }
  .signal-path { stroke-dasharray: 8 10; animation: dashFlow 1.6s linear infinite; }
  .route-path { stroke-dasharray: 12 14; animation: dashFlow 2.8s linear infinite; }
  .scan-band { animation: scanY 3s linear infinite; }
  .pulse-core { animation: pulseCore 1.6s ease-in-out infinite; transform-origin: center; }
  @keyframes dashFlow { to { stroke-dashoffset: -44; } }
  @keyframes scanY { from { transform: translateY(-220px); opacity: .15; } 45% { opacity: .8; } to { transform: translateY(260px); opacity: .15; } }
  @keyframes pulseCore { 0%,100% { opacity: .55; transform: scale(1); } 50% { opacity: 1; transform: scale(1.12); } }
</style>

{% set total = ventas | sum(attribute='monto') %}
{% set ticket_promedio = total / (ventas | length) %}
{% set cumplimiento = total / meta_global * 100 %}

<div class="bi-hero">
  <h1>{{ empresa }}</h1>
  <p>{{ periodo }} - experimento: Jinja2 genera una escena SVG animada desde JSON.</p>
  <p>Meta global: {{ meta_global | currency }} | Cumplimiento: <strong>{{ cumplimiento | round(1) | percent }}</strong></p>
</div>

<div class="bi-grid">
  <div class="bi-kpi"><small>Venta total</small><strong>{{ total | currency }}</strong></div>
  <div class="bi-kpi"><small>Ticket promedio</small><strong>{{ ticket_promedio | currency }}</strong></div>
  <div class="bi-kpi"><small>Alertas criticas</small><strong>{{ alertas | length }}</strong></div>
</div>

<h2>Tabla dinamica: region x producto</h2>
<table class="pivot">
  <thead>
    <tr>
      <th>Region</th>
      {% for producto in productos %}<th>{{ producto }}</th>{% endfor %}
      <th>Total</th>
    </tr>
  </thead>
  <tbody>
    {% for region in regiones %}
    {% set total_region = ventas | selectattr("region", "equalto", region) | sum(attribute="monto") %}
    <tr>
      <td>{{ region }}</td>
      {% for producto in productos %}
      {% set celda = ventas
        | selectattr("region", "equalto", region)
        | selectattr("producto", "equalto", producto)
        | sum(attribute="monto") %}
      <td class="{% if celda >= 40000 %}heat-high{% elif celda >= 20000 %}heat-mid{% else %}heat-low{% endif %}">
        {{ celda | currency }}
      </td>
      {% endfor %}
      <td><strong>{{ total_region | currency }}</strong></td>
    </tr>
    {% endfor %}
  </tbody>
</table>

{% set bar_chart = {
  "type": "bar",
  "title": "Ventas por region",
  "labels": regiones,
  "datasets": [{
    "label": "Total vendido",
    "data": totales_region,
    "backgroundColor": ["#2563eb", "#16a34a", "#f59e0b", "#dc2626"]
  }]
} %}
{{ bar_chart | chart }}

{% set donut_chart = {
  "type": "doughnut",
  "title": "Mix por producto",
  "labels": productos,
  "datasets": [{
    "label": "Participacion",
    "data": totales_producto,
    "backgroundColor": ["#7c3aed", "#0891b2", "#f97316", "#84cc16"]
  }]
} %}
{{ donut_chart | chart }}

<div class="svg-panel">
  <h2>Escena SVG: red logistica viva</h2>
  <svg viewBox="0 0 760 340" width="100%" height="340" role="img" aria-label="Red logistica animada">
    <defs>
      <radialGradient id="coreGlow" cx="50%" cy="50%" r="60%">
        <stop offset="0%" stop-color="#67e8f9" stop-opacity=".95"></stop>
        <stop offset="55%" stop-color="#38bdf8" stop-opacity=".28"></stop>
        <stop offset="100%" stop-color="#020617" stop-opacity="0"></stop>
      </radialGradient>
      <linearGradient id="routeGlow" x1="0" x2="1">
        <stop offset="0%" stop-color="#22d3ee"></stop>
        <stop offset="50%" stop-color="#a78bfa"></stop>
        <stop offset="100%" stop-color="#f472b6"></stop>
      </linearGradient>
      <filter id="softGlow">
        <feGaussianBlur stdDeviation="4" result="blur"></feGaussianBlur>
        <feMerge><feMergeNode in="blur"></feMergeNode><feMergeNode in="SourceGraphic"></feMergeNode></feMerge>
      </filter>
      {% for ruta in rutas %}
      <path id="ruta{{ loop.index }}" d="{{ ruta.path }}"></path>
      {% endfor %}
    </defs>

    <rect width="760" height="340" rx="16" fill="#020617"></rect>
    <g opacity=".22">
      {% for x in range(40, 760, 40) %}<line x1="{{ x }}" y1="0" x2="{{ x }}" y2="340" stroke="#334155" stroke-width=".7"></line>{% endfor %}
      {% for y in range(35, 340, 35) %}<line x1="0" y1="{{ y }}" x2="760" y2="{{ y }}" stroke="#334155" stroke-width=".7"></line>{% endfor %}
    </g>

    <rect class="scan-band" x="0" y="0" width="760" height="54" fill="#22d3ee" opacity=".12"></rect>

    {% for ruta in rutas %}
    <use href="#ruta{{ loop.index }}" fill="none" stroke="#1e293b" stroke-width="{{ ruta.width + 8 }}" opacity=".65"></use>
    <use href="#ruta{{ loop.index }}" class="route-path" fill="none" stroke="url(#routeGlow)" stroke-width="{{ ruta.width }}" stroke-linecap="round" opacity=".9"></use>
    <circle r="{{ ruta.packet }}" fill="{{ ruta.color }}" filter="url(#softGlow)">
      <animateMotion dur="{{ ruta.dur }}s" repeatCount="indefinite" begin="{{ ruta.delay }}s" rotate="auto">
        <mpath href="#ruta{{ loop.index }}"></mpath>
      </animateMotion>
      <animate attributeName="opacity" values=".25;1;.25" dur="{{ ruta.dur }}s" repeatCount="indefinite"></animate>
    </circle>
    {% endfor %}

    <g transform="translate(380 170)">
      <circle r="90" fill="url(#coreGlow)" opacity=".7"></circle>
      <circle r="54" fill="#0f172a" stroke="#67e8f9" stroke-width="2" opacity=".92"></circle>
      <circle class="pulse-core" r="22" fill="#22d3ee" filter="url(#softGlow)"></circle>
      <g>
        <circle r="118" fill="none" stroke="#38bdf8" stroke-width="1" stroke-dasharray="6 12" opacity=".65"></circle>
        <circle r="142" fill="none" stroke="#a78bfa" stroke-width="1" stroke-dasharray="3 14" opacity=".5"></circle>
        <animateTransform attributeName="transform" type="rotate" from="0" to="360" dur="18s" repeatCount="indefinite"></animateTransform>
      </g>
      <text y="6" fill="#e0f2fe" text-anchor="middle" font-size="14" font-weight="700">CORE</text>
      <text y="28" fill="#94a3b8" text-anchor="middle" font-size="10">{{ cumplimiento | round(1) }}% sync</text>
    </g>

    {% for hub in hubs %}
    <g transform="translate({{ hub.x }} {{ hub.y }})">
      <circle r="{{ 18 + hub.score / 8 }}" fill="{{ hub.color }}" opacity=".2">
        <animate attributeName="r" values="{{ 18 + hub.score / 8 }};{{ 34 + hub.score / 8 }};{{ 18 + hub.score / 8 }}" dur="{{ hub.pulse }}s" repeatCount="indefinite"></animate>
        <animate attributeName="opacity" values=".18;.02;.18" dur="{{ hub.pulse }}s" repeatCount="indefinite"></animate>
      </circle>
      <animateTransform attributeName="transform" type="translate"
        values="{{ hub.x }} {{ hub.y }}; {{ hub.x + 2 }} {{ hub.y - 1 }}; {{ hub.x - 1 }} {{ hub.y + 2 }}; {{ hub.x }} {{ hub.y }}"
        dur="{{ hub.pulse + 2 }}s" repeatCount="indefinite"></animateTransform>
      <rect x="-34" y="-18" width="68" height="36" rx="10" fill="#0f172a" stroke="{{ hub.color }}" stroke-width="2"></rect>
      <circle cx="-20" cy="0" r="6" fill="{{ hub.color }}">
        <animate attributeName="opacity" values=".35;1;.35" dur="{{ hub.pulse }}s" repeatCount="indefinite"></animate>
      </circle>
      <text x="0" y="4" fill="#e5e7eb" text-anchor="middle" font-size="11" font-weight="700">{{ hub.nombre }}</text>
      <text x="0" y="35" fill="#94a3b8" text-anchor="middle" font-size="10">{{ hub.score }} pts</text>
    </g>
    {% endfor %}

    {% for alerta in alertas %}
    <g transform="translate({{ alerta.x }} {{ alerta.y }})">
      <polygon points="0,-14 14,12 -14,12" fill="{{ alerta.color }}" opacity=".9"></polygon>
      <text x="0" y="6" fill="#111827" text-anchor="middle" font-size="13" font-weight="900">!</text>
      <animate attributeName="opacity" values=".55;1;.55" dur="{{ alerta.beat }}s" repeatCount="indefinite"></animate>
    </g>
    {% endfor %}
  </svg>
  <p class="scene-caption">Todo lo que se mueve aqui viene del JSON: hubs, rutas, alertas, colores, tiempos y trayectorias SVG.</p>
</div>

<h2>Alertas generadas desde datos</h2>
{% for alerta in alertas %}
<div class="card">
  <span class="{{ alerta.riesgo | badge }}">{{ alerta.riesgo | percent }}</span>
  <strong>{{ alerta.area }}</strong> - {{ alerta.mensaje }}
</div>
{% endfor %}

<div class="svg-panel video-stage">
  <h2>Simulacion de video: camara de operaciones</h2>
  <svg viewBox="0 0 760 300" width="100%" height="300" role="img" aria-label="Video simulado con SVG">
    <defs>
      <clipPath id="cinemaClip"><rect x="18" y="18" width="724" height="210" rx="16"></rect></clipPath>
      <linearGradient id="sky" x1="0" x2="0" y1="0" y2="1">
        <stop offset="0%" stop-color="#0f172a"></stop>
        <stop offset="100%" stop-color="#1e293b"></stop>
      </linearGradient>
      <filter id="blurGlow"><feGaussianBlur stdDeviation="3"></feGaussianBlur></filter>
      {% for toma in tomas %}
      <path id="toma{{ loop.index }}" d="{{ toma.path }}"></path>
      {% endfor %}
    </defs>

    <rect width="760" height="300" rx="18" fill="#020617"></rect>
    <g clip-path="url(#cinemaClip)">
      <rect x="18" y="18" width="724" height="210" fill="url(#sky)"></rect>
      <circle cx="610" cy="65" r="34" fill="#fde68a" opacity=".88">
        <animate attributeName="r" values="30;38;30" dur="4s" repeatCount="indefinite"></animate>
      </circle>
      {% for nube in nubes %}
      <g opacity=".55">
        <ellipse cx="{{ nube.x }}" cy="{{ nube.y }}" rx="{{ nube.rx }}" ry="12" fill="#94a3b8"></ellipse>
        <ellipse cx="{{ nube.x + 22 }}" cy="{{ nube.y - 5 }}" rx="{{ nube.rx / 1.5 }}" ry="10" fill="#cbd5e1"></ellipse>
        <animateTransform attributeName="transform" type="translate" values="0 0; {{ nube.move }} 0; 0 0" dur="{{ nube.dur }}s" repeatCount="indefinite"></animateTransform>
      </g>
      {% endfor %}

      <path d="M18 178 C130 128, 220 220, 340 168 S560 118,742 170 L742 228 L18 228 Z" fill="#064e3b"></path>
      <path d="M18 202 C150 172, 242 238, 392 190 S610 158,742 196" fill="none" stroke="#a7f3d0" stroke-width="7" opacity=".7"></path>

      {% for toma in tomas %}
      <use href="#toma{{ loop.index }}" class="signal-path" fill="none" stroke="{{ toma.color }}" stroke-width="3" opacity=".75"></use>
      <g>
        <rect x="-18" y="-9" width="36" height="18" rx="5" fill="{{ toma.color }}" stroke="#e0f2fe" stroke-width="1.5"></rect>
        <circle cx="11" cy="0" r="3" fill="#020617"></circle>
        <animateMotion dur="{{ toma.dur }}s" repeatCount="indefinite" rotate="auto" begin="{{ toma.delay }}s">
          <mpath href="#toma{{ loop.index }}"></mpath>
        </animateMotion>
      </g>
      {% endfor %}

      <rect class="scan-band" x="18" y="18" width="724" height="42" fill="#67e8f9" opacity=".16"></rect>
      <g transform="translate(520 92)">
        <rect x="-70" y="-38" width="140" height="76" rx="12" fill="#020617" opacity=".72" stroke="#38bdf8"></rect>
        <circle r="24" fill="none" stroke="#38bdf8" stroke-width="3"></circle>
        <line x1="-34" y1="0" x2="34" y2="0" stroke="#38bdf8"></line>
        <line x1="0" y1="-34" x2="0" y2="34" stroke="#38bdf8"></line>
        <animateTransform attributeName="transform" type="translate" values="520 92; 242 162; 604 176; 520 92" dur="9s" repeatCount="indefinite"></animateTransform>
      </g>
    </g>

    <rect x="18" y="18" width="724" height="210" rx="16" fill="none" stroke="#475569"></rect>
    <text x="36" y="252" fill="#e2e8f0" font-size="13" font-weight="700">Timeline generado por plantilla</text>
    {% for frame in frames %}
    <g transform="translate({{ 36 + loop.index0 * 108 }} 264)">
      <rect width="86" height="18" rx="9" fill="#1e293b" stroke="{{ frame.color }}"></rect>
      <rect width="{{ frame.width }}" height="18" rx="9" fill="{{ frame.color }}" opacity=".72">
        <animate attributeName="width" values="8;{{ frame.width }};8" dur="{{ frame.dur }}s" repeatCount="indefinite" begin="{{ loop.index0 * 0.18 }}s"></animate>
      </rect>
      <text x="43" y="13" text-anchor="middle" fill="#f8fafc" font-size="10">{{ frame.label }}</text>
    </g>
    {% endfor %}
  </svg>
  <p class="scene-caption">No hay archivo de video. Jinja2 escribe SVG + SMIL: rutas, camara, timeline, scan y vehiculos animados.</p>
</div>

{% if _role == 'admin' %}
<div class="admin-section">
  <h2>Panel secreto de administracion</h2>
  <p>Margen neto real: <strong>{{ margen_neto | percent }}</strong></p>
  <p>Riesgo financiero estimado: <strong>{{ riesgo_financiero | percent }}</strong></p>
  <p>Decision sugerida: {{ decision_admin }}</p>
</div>
{% else %}
<div class="card">
  <span class="badge-warning">Modo viewer</span>
  Cambia a admin para revelar el panel confidencial.
</div>
{% endif %}
""",
        "data": json.dumps({
            "empresa": "Andes Tech Command Center",
            "periodo": "Junio 2026",
            "meta_global": 310000,
            "regiones": ["Lima", "Norte", "Sur", "Centro"],
            "productos": ["Laptop", "Tablet", "Monitor", "Software"],
            "ventas": [
                {"region": "Lima", "producto": "Laptop", "monto": 52000},
                {"region": "Lima", "producto": "Tablet", "monto": 24000},
                {"region": "Lima", "producto": "Monitor", "monto": 31000},
                {"region": "Lima", "producto": "Software", "monto": 45000},
                {"region": "Norte", "producto": "Laptop", "monto": 28000},
                {"region": "Norte", "producto": "Tablet", "monto": 18000},
                {"region": "Norte", "producto": "Monitor", "monto": 22000},
                {"region": "Norte", "producto": "Software", "monto": 34000},
                {"region": "Sur", "producto": "Laptop", "monto": 19000},
                {"region": "Sur", "producto": "Tablet", "monto": 12000},
                {"region": "Sur", "producto": "Monitor", "monto": 16000},
                {"region": "Sur", "producto": "Software", "monto": 21000},
                {"region": "Centro", "producto": "Laptop", "monto": 37000},
                {"region": "Centro", "producto": "Tablet", "monto": 23000},
                {"region": "Centro", "producto": "Monitor", "monto": 26000},
                {"region": "Centro", "producto": "Software", "monto": 39000}
            ],
            "totales_region": [152000, 102000, 68000, 125000],
            "totales_producto": [136000, 77000, 95000, 139000],
            "hubs": [
                {"nombre": "Lima", "x": 115, "y": 76, "score": 96, "color": "#22c55e", "pulse": 1.8},
                {"nombre": "Norte", "x": 226, "y": 254, "score": 72, "color": "#38bdf8", "pulse": 2.2},
                {"nombre": "Sur", "x": 620, "y": 248, "score": 41, "color": "#f59e0b", "pulse": 1.5},
                {"nombre": "Centro", "x": 650, "y": 78, "score": 84, "color": "#a78bfa", "pulse": 2.6}
            ],
            "rutas": [
                {"path": "M115 76 C205 42, 295 52, 380 170", "width": 4, "packet": 7, "dur": 4.6, "delay": 0, "color": "#22d3ee"},
                {"path": "M226 254 C276 222, 312 210, 380 170", "width": 5, "packet": 8, "dur": 3.9, "delay": .35, "color": "#38bdf8"},
                {"path": "M620 248 C548 236, 488 214, 380 170", "width": 4, "packet": 7, "dur": 5.1, "delay": .7, "color": "#f472b6"},
                {"path": "M650 78 C568 74, 478 92, 380 170", "width": 5, "packet": 8, "dur": 4.2, "delay": 1.1, "color": "#a78bfa"},
                {"path": "M115 76 C188 150, 266 192, 226 254", "width": 3, "packet": 5, "dur": 6.4, "delay": .2, "color": "#84cc16"},
                {"path": "M650 78 C680 142, 690 208, 620 248", "width": 3, "packet": 5, "dur": 5.8, "delay": .9, "color": "#f59e0b"}
            ],
            "alertas": [
                {"area": "Sur", "riesgo": 28, "mensaje": "Baja conversion en canal corporativo", "x": 588, "y": 214, "color": "#f97316", "beat": 1.2},
                {"area": "Norte", "riesgo": 64, "mensaje": "Stock ajustado para monitores", "x": 258, "y": 224, "color": "#facc15", "beat": 1.7},
                {"area": "Lima", "riesgo": 87, "mensaje": "Campania de software supera meta", "x": 154, "y": 108, "color": "#22c55e", "beat": 2.1}
            ],
            "nubes": [
                {"x": 95, "y": 58, "rx": 38, "move": 34, "dur": 12},
                {"x": 332, "y": 72, "rx": 30, "move": -28, "dur": 10},
                {"x": 578, "y": 48, "rx": 42, "move": 42, "dur": 15}
            ],
            "tomas": [
                {"path": "M88 190 C190 118, 310 128, 430 172 S620 226, 690 142", "color": "#22d3ee", "dur": 7.6, "delay": 0},
                {"path": "M70 152 C190 216, 306 210, 406 150 S584 92, 704 172", "color": "#a78bfa", "dur": 8.8, "delay": 1.4},
                {"path": "M112 214 C224 172, 346 196, 468 134 S604 128, 676 202", "color": "#f472b6", "dur": 6.9, "delay": 2.2}
            ],
            "frames": [
                {"label": "scan", "width": 70, "dur": 2.2, "color": "#22d3ee"},
                {"label": "ruta", "width": 58, "dur": 2.8, "color": "#38bdf8"},
                {"label": "hub", "width": 76, "dur": 1.9, "color": "#818cf8"},
                {"label": "alerta", "width": 50, "dur": 2.4, "color": "#f59e0b"},
                {"label": "camara", "width": 82, "dur": 3.1, "color": "#f472b6"},
                {"label": "cierre", "width": 64, "dur": 2.6, "color": "#22c55e"}
            ],
            "margen_neto": 24.8,
            "riesgo_financiero": 31.5,
            "decision_admin": "Reasignar presupuesto desde Sur hacia campanias de software en Lima y Centro."
        }, indent=2)
    }
}


# ── Helpers ──────────────────────────────────────────────────────────────────

def next_run_from_frequency(frequency: str) -> str:
    now = datetime.utcnow()
    if frequency == "daily":
        dt = now + timedelta(days=1)
    elif frequency == "weekly":
        dt = now + timedelta(weeks=1)
    else:  # monthly
        # same day next month (approx)
        month = now.month + 1 if now.month < 12 else 1
        year  = now.year if now.month < 12 else now.year + 1
        dt = now.replace(year=year, month=month)
    return dt.isoformat()

def current_role():
    return session.get("role", "viewer")

def current_user():
    return session.get("username", "anónimo")


# ── Routes ───────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template(
        "index.html",
        examples=EXAMPLE_TEMPLATES,
        role=current_role(),
        username=current_user(),
    )


@app.route("/set-role", methods=["POST"])
def set_role():
    body = request.get_json()
    role = body.get("role", "viewer")
    username = body.get("username", "").strip() or ("Admin" if role == "admin" else "Visitante")
    if role not in ("admin", "viewer"):
        return jsonify({"error": "Rol inválido"}), 400
    session["role"] = role
    session["username"] = username
    return jsonify({"role": role, "username": username})


@app.route("/preview", methods=["POST"])
def preview():
    body = request.get_json()
    template_src = body.get("template", "")
    data_src = body.get("data", "{}")

    try:
        context = json.loads(data_src)
    except json.JSONDecodeError as e:
        return jsonify({"error": f"JSON inválido: {e}"}), 400

    # Inject role/user — cannot be overridden by template data
    context["_role"] = current_role()
    context["_user"] = current_user()

    try:
        tmpl = user_env.from_string(template_src)
        rendered = tmpl.render(**context)
        return jsonify({"html": rendered})
    except TemplateSyntaxError as e:
        return jsonify({"error": f"Error de sintaxis Jinja2 (línea {e.lineno}): {e.message}"}), 400
    except UndefinedError as e:
        return jsonify({"error": f"Variable no definida: {e}"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/examples/<key>")
def get_example(key):
    ex = EXAMPLE_TEMPLATES.get(key)
    if not ex:
        return jsonify({"error": "Ejemplo no encontrado"}), 404
    return jsonify(ex)


# ── Export: Excel ─────────────────────────────────────────────────────────────

@app.route("/export/excel", methods=["POST"])
def export_excel():
    body = request.get_json()
    data_src  = body.get("data", "{}")
    filename  = body.get("filename", "reporte").strip() or "reporte"

    try:
        context = json.loads(data_src)
    except json.JSONDecodeError as e:
        return jsonify({"error": f"JSON inválido: {e}"}), 400

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="7C6AF7")
    header_align = Alignment(horizontal="center")

    sheets_created = 0

    for key, value in context.items():
        if isinstance(value, list) and len(value) > 0 and isinstance(value[0], dict):
            ws = wb.create_sheet(title=str(key)[:31])
            headers = list(value[0].keys())

            # Header row
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font  = header_font
                cell.fill  = header_fill
                cell.alignment = header_align

            # Data rows
            for row_idx, row in enumerate(value, start=2):
                for col_idx, header in enumerate(headers, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(header, ""))

            # Auto column width
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

            sheets_created += 1

    # Fallback: scalar values sheet
    if sheets_created == 0:
        ws = wb.create_sheet(title="datos")
        ws.cell(row=1, column=1, value="Clave").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=2, value="Valor").font = header_font
        ws.cell(row=1, column=2).fill = header_fill
        row_idx = 2
        for k, v in context.items():
            if not isinstance(v, (list, dict)):
                ws.cell(row=row_idx, column=1, value=k)
                ws.cell(row=row_idx, column=2, value=str(v))
                row_idx += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{filename}.xlsx",
    )


# ── API: CRUD de plantillas guardadas ────────────────────────────────────────

@app.route("/api/templates", methods=["GET"])
def list_templates():
    conn = get_db()
    role = current_role()
    if role == "admin":
        rows = conn.execute(
            "SELECT id, name, version, visibility, updated_at FROM saved_templates ORDER BY updated_at DESC"
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT id, name, version, visibility, updated_at FROM saved_templates WHERE visibility='public' ORDER BY updated_at DESC"
        ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/templates", methods=["POST"])
def save_template():
    body = request.get_json()
    name       = body.get("name", "").strip()
    visibility = body.get("visibility", "public")
    if not name:
        return jsonify({"error": "El nombre es obligatorio"}), 400
    if visibility not in ("public", "private"):
        visibility = "public"

    conn = get_db()
    cur = conn.execute(
        "INSERT INTO saved_templates (name, template, data, visibility) VALUES (?, ?, ?, ?)",
        (name, body.get("template", ""), body.get("data", "{}"), visibility),
    )
    conn.commit()
    row = conn.execute(
        "SELECT id, name, version, visibility, updated_at FROM saved_templates WHERE id = ?",
        (cur.lastrowid,),
    ).fetchone()
    conn.close()
    return jsonify(dict(row)), 201


@app.route("/api/templates/<int:tid>", methods=["GET"])
def get_template(tid):
    conn = get_db()
    row = conn.execute("SELECT * FROM saved_templates WHERE id = ?", (tid,)).fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "Plantilla no encontrada"}), 404
    r = dict(row)
    if r.get("visibility") == "private" and current_role() != "admin":
        return jsonify({"error": "Acceso denegado: plantilla privada"}), 403
    return jsonify(r)


@app.route("/api/templates/<int:tid>", methods=["PUT"])
def update_template(tid):
    conn = get_db()
    existing = conn.execute("SELECT * FROM saved_templates WHERE id = ?", (tid,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({"error": "Plantilla no encontrada"}), 404

    body       = request.get_json()
    name       = body.get("name", existing["name"]).strip()
    visibility = body.get("visibility", existing["visibility"])
    if not name:
        return jsonify({"error": "El nombre es obligatorio"}), 400
    if visibility not in ("public", "private"):
        visibility = existing["visibility"]

    conn.execute(
        """UPDATE saved_templates
           SET name = ?, template = ?, data = ?, visibility = ?,
               version = version + 1, updated_at = ?
           WHERE id = ?""",
        (name, body.get("template", existing["template"]),
         body.get("data", existing["data"]), visibility,
         datetime.utcnow().isoformat(), tid),
    )
    conn.commit()
    row = conn.execute(
        "SELECT id, name, version, visibility, updated_at FROM saved_templates WHERE id = ?", (tid,)
    ).fetchone()
    conn.close()
    return jsonify(dict(row))


@app.route("/api/templates/<int:tid>", methods=["DELETE"])
def delete_template(tid):
    conn = get_db()
    cur = conn.execute("DELETE FROM saved_templates WHERE id = ?", (tid,))
    conn.commit()
    conn.close()
    if cur.rowcount == 0:
        return jsonify({"error": "Plantilla no encontrada"}), 404
    return jsonify({"ok": True})


# ── API: Scheduling ───────────────────────────────────────────────────────────

@app.route("/api/schedules", methods=["GET"])
def list_schedules():
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM scheduled_reports ORDER BY next_run ASC"
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/schedules", methods=["POST"])
def create_schedule():
    body      = request.get_json()
    name      = body.get("name", "").strip()
    tid       = body.get("template_id")
    frequency = body.get("frequency", "daily")

    if not name:
        return jsonify({"error": "El nombre es obligatorio"}), 400
    if frequency not in ("daily", "weekly", "monthly"):
        return jsonify({"error": "Frecuencia inválida"}), 400

    next_run = next_run_from_frequency(frequency)
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO scheduled_reports (name, template_id, frequency, next_run) VALUES (?, ?, ?, ?)",
        (name, tid, frequency, next_run),
    )
    conn.commit()
    row = conn.execute(
        "SELECT * FROM scheduled_reports WHERE id = ?", (cur.lastrowid,)
    ).fetchone()
    conn.close()
    return jsonify(dict(row)), 201


@app.route("/api/schedules/<int:sid>", methods=["DELETE"])
def delete_schedule(sid):
    conn = get_db()
    cur = conn.execute("DELETE FROM scheduled_reports WHERE id = ?", (sid,))
    conn.commit()
    conn.close()
    if cur.rowcount == 0:
        return jsonify({"error": "Programación no encontrada"}), 404
    return jsonify({"ok": True})


@app.route("/api/schedules/<int:sid>/toggle", methods=["POST"])
def toggle_schedule(sid):
    conn = get_db()
    row = conn.execute("SELECT active FROM scheduled_reports WHERE id = ?", (sid,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "No encontrado"}), 404
    new_active = 0 if row["active"] else 1
    conn.execute("UPDATE scheduled_reports SET active = ? WHERE id = ?", (new_active, sid))
    conn.commit()
    row = conn.execute("SELECT * FROM scheduled_reports WHERE id = ?", (sid,)).fetchone()
    conn.close()
    return jsonify(dict(row))


@app.route("/api/schedules/<int:sid>/run", methods=["POST"])
def run_schedule(sid):
    conn = get_db()
    sched = conn.execute("SELECT * FROM scheduled_reports WHERE id = ?", (sid,)).fetchone()
    if not sched:
        conn.close()
        return jsonify({"error": "No encontrado"}), 404

    s = dict(sched)
    next_run = next_run_from_frequency(s["frequency"])
    conn.execute(
        "UPDATE scheduled_reports SET last_run = ?, next_run = ? WHERE id = ?",
        (datetime.utcnow().isoformat(), next_run, sid),
    )
    conn.execute(
        "INSERT INTO report_runs (schedule_id, schedule_name, status) VALUES (?, ?, ?)",
        (sid, s["name"], "ok"),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM scheduled_reports WHERE id = ?", (sid,)).fetchone()
    runs = conn.execute(
        "SELECT * FROM report_runs WHERE schedule_id = ? ORDER BY ran_at DESC LIMIT 5", (sid,)
    ).fetchall()
    conn.close()
    return jsonify({"schedule": dict(row), "last_runs": [dict(r) for r in runs]})


@app.route("/api/schedules/due", methods=["GET"])
def due_schedules():
    now = datetime.utcnow().isoformat()
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM scheduled_reports WHERE active = 1 AND next_run <= ?", (now,)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route("/api/schedules/<int:sid>/runs", methods=["GET"])
def schedule_runs(sid):
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM report_runs WHERE schedule_id = ? ORDER BY ran_at DESC LIMIT 20", (sid,)
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


if __name__ == "__main__":
    app.run(debug=True, port=5000)
